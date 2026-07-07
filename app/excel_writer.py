from __future__ import annotations

from copy import copy
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.parser import FillInstruction, SequenceItem


def fill_workbook(
    input_path: str,
    output_path: str,
    sheet_name: str,
    instruction: FillInstruction,
) -> None:
    workbook = load_workbook(input_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    _apply_instruction(sheet, instruction)
    workbook.save(output_path)


def _apply_instruction(sheet: Worksheet, instruction: FillInstruction) -> None:
    current_row = instruction.start_row

    for item in instruction.items:
        if item.inserted:
            sheet.insert_rows(current_row)
            _copy_row_style(sheet, current_row + 1, current_row)

        sheet[f"B{current_row}"] = _compose_value(instruction.prefix, item, instruction.suffix)
        current_row += 1


def _compose_value(prefix: str, item: SequenceItem, suffix: str) -> str:
    return f"{prefix}{item.value}{suffix}"


def _copy_row_style(sheet: Worksheet, source_row: int, target_row: int) -> None:
    for source_cell, target_cell in zip(
        sheet[source_row],
        sheet[target_row],
        strict=False,
    ):
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)

        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    _copy_dimensions(sheet, source_row, target_row)


def _copy_dimensions(sheet: Worksheet, source_row: int, target_row: int) -> None:
    source_dimensions = sheet.row_dimensions.get(source_row)
    if source_dimensions is not None:
        sheet.row_dimensions[target_row].height = source_dimensions.height

    merged_ranges = list(_merged_ranges_touching_row(sheet.merged_cells.ranges, source_row))
    for merged_range in merged_ranges:
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            offset = target_row - source_row
            sheet.merge_cells(
                start_row=merged_range.min_row + offset,
                start_column=merged_range.min_col,
                end_row=merged_range.max_row + offset,
                end_column=merged_range.max_col,
            )


def _merged_ranges_touching_row(ranges: Iterable, row: int):
    for merged_range in ranges:
        if merged_range.min_row <= row <= merged_range.max_row:
            yield merged_range
